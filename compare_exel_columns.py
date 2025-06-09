import pandas as pd
import warnings
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# Suppress openpyxl style warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

def preprocess_excel(file_path):
    """Preprocess Excel file to reset problematic styles."""
    try:
        # Load workbook
        wb = load_workbook(file_path)
        
        # Iterate through all worksheets
        for sheet in wb:
            for row in sheet.iter_rows():
                for cell in row:
                    # Reset fill to default if it exists
                    if cell.fill:
                        cell.fill = PatternFill()  # Default empty fill
                        
        # Save to a temporary file
        temp_file = file_path.replace('.xlsx', '_temp.xlsx')
        wb.save(temp_file)
        return temp_file
    except Exception as e:
        print(f"Error preprocessing {file_path}: {str(e)}")
        return file_path

def normalize_column_name(name):
    """Normalize column name by converting to lowercase and replacing spaces with underscores."""
    return name.lower().replace(' ', '_')

def compare_excel_columns(file1_path, file2_path, column_name, output_file="mismatches.xlsx", sheet_name1=0, sheet_name2=0):
    try:
        # Preprocess files to handle style issues
        file1_path = preprocess_excel(file1_path)
        file2_path = preprocess_excel(file2_path)
        
        # Read the Excel files
        df1 = pd.read_excel(file1_path, sheet_name=sheet_name1, engine='openpyxl')
        df2 = pd.read_excel(file2_path, sheet_name=sheet_name2, engine='openpyxl')
        
        # Normalize column names for case-insensitive and space/underscore matching
        df1_columns = {normalize_column_name(col): col for col in df1.columns}
        df2_columns = {normalize_column_name(col): col for col in df2.columns}
        column_name_normalized = normalize_column_name(column_name)
        
        # Check if the column exists in both files
        if column_name_normalized not in df1_columns or column_name_normalized not in df2_columns:
            print(f"Error: Column '{column_name}' (case-insensitive, ignoring spaces/underscores) not found in one or both files.")
            return
        
        # Get the original column names
        col1 = df1_columns[column_name_normalized]
        col2 = df2_columns[column_name_normalized]
        
        # Rename the column to a consistent name for merging
        df1 = df1.rename(columns={col1: 'normalized_column'})
        df2 = df2.rename(columns={col2: 'normalized_column'})
        
        # Merge dataframes on the normalized column to find mismatches
        merged = pd.merge(df1, df2, on='normalized_column', how='outer', indicator=True)
        
        # Filter rows that are only in one file (mismatches)
        mismatches = merged[merged['_merge'] != 'both']
        
        if mismatches.empty:
            print("No mismatches found in the specified column.")
        else:
            # Prepare mismatches for saving
            mismatches_output = mismatches.drop('_merge', axis=1).rename(columns={'normalized_column': 'PROMOTION_CODE'})
            # Add a column to indicate the source file
            mismatches_output['Source'] = mismatches['_merge'].map({'left_only': 'File 1', 'right_only': 'File 2'})
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else '.', exist_ok=True)
            # Save to Excel
            mismatches_output.to_excel(output_file, index=False, engine='openpyxl')
            print(f"Mismatched rows saved to '{output_file}'.")
            print(f"Absolute path: {os.path.abspath(output_file)}")
                
    except FileNotFoundError:
        print("Error: One or both files not found.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
    finally:
        # Clean up temporary files if created
        for temp_file in [file1_path.replace('.xlsx', '_temp.xlsx'), file2_path.replace('.xlsx', '_temp.xlsx')]:
            if os.path.exists(temp_file):
                os.remove(temp_file)

# Example usage
if __name__ == "__main__":
    file1 = "file1.xlsx"  # Replace with your first Excel file path
    file2 = "file2.xlsx"  # Replace with your second Excel file path
    column_to_compare = "PROMOTION_CODE"  # Column name, will match 'Promotion Code'
    output_file = "mismatches.xlsx"  # Output file for mismatches
    
    compare_excel_columns(file1, file2, column_to_compare, output_file)
